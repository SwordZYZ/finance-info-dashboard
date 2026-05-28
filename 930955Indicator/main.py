# /// script
# dependencies = [
#   "requests",
#   "pandas",
#   "xlrd",
#   "openpyxl",
# ]
# ///
import os
import io
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook

# 1. 定义配置
url = "https://oss-ch.csindex.com.cn/static/html/csindex/public/uploads/file/autofile/indicator/930955indicator.xls"
output_file = "930955Indicator.xlsx"
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
}

try:
    print("正在获取最新数据...")
    # 2. 抓取中证官网数据
    response = requests.get(url, headers=headers, timeout=15)
    response.raise_for_status()
    
    # 读取下载的 .xls 文件
    df_web = pd.read_excel(io.BytesIO(response.content))
    
    if df_web.empty:
        print("未能在网上获取到有效数据。")
        exit()
        
    # 获取第一行数据的具体值（转化为列表，保持顺序与表头一致）
    headers_list = df_web.columns.tolist()
    latest_row_values = df_web.iloc[0].tolist()
    
    current_date = latest_row_values[0] # 获取当前抓取数据的日期
    print(f"成功获取到 {current_date} 的数据。")

    # 3. 写入/插入到本地的 Excel 文件
    if os.path.exists(output_file):
        print(f"发现本地文件 {output_file}，正在插入到第二行...")
        
        # 加载已有工作簿
        wb = load_workbook(output_file)
        ws = wb.active
        
        # 检查是否已经存在该日期的数据，防止重复插入
        date_exists = False
        # 遍历第二行及之后第一列的日期（假设第一列是日期）
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row=row, column=1).value) == str(current_date):
                date_exists = True
                break
        
        if date_exists:
            print(f"提示：本地文件中已存在日期为 {current_date} 的数据，跳过插入，避免重复。")
        else:
            # 在第二行（idx=2）位置插入一个空行
            ws.insert_rows(idx=2, amount=1)
            
            # 将最新的数据写入到刚刚插入的第二行中
            for col_idx, value in enumerate(latest_row_values, start=1):
                ws.cell(row=2, column=col_idx, value=value)
                
            wb.save(output_file)
            print("数据成功插入到第二行！")
            
    else:
        print(f"本地未找到 {output_file}，将自动创建并写入数据...")
        # 如果文件不存在，直接用 pandas 创建一个全新的文件
        wb = Workbook()
        ws = wb.active
        # 写入表头
        ws.append(headers_list)
        # 写入第一行数据
        ws.append(latest_row_values)
        wb.save(output_file)
        print(f"新文件 {output_file} 创建并写入成功！")

except Exception as e:
    print(f"执行过程中发生错误: {e}")

# 顺便把整个更新后的 Excel 转换为网页最喜欢的 JSON 格式
# 注意：为了让图表从左到右按时间正序排列，我们需要把倒序的 Excel 翻转一下 (.iloc[::-1])
df_local = pd.read_excel(output_file)

# 2使用 .head(20) 只获取最新的 20 条数据
# 然后使用 .iloc[::-1] 将其翻转，使得图表在网页上从左到右按时间正序排列
df_json = df_local.head(20).iloc[::-1]

# 只保留网页图表需要的列，减少文件体积：日期、股息率1、股息率2
df_json = df_json[['日期Date', '股息率1（总股本）D/P1', '股息率2（计算用股本）D/P2']]

# 导出为 json 文件
df_json.to_json("data.json", orient="records", force_ascii=False)
print("本地 data.json 数据同步更新成功！")